"""Shared safe CSV and XLSX serializers for report data."""

import csv
import io
import re
from datetime import date, datetime
from decimal import Decimal

from openpyxl import Workbook
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

from backend.reports.registry import ReportColumn, ReportData

CSV_CONTENT_TYPE = "text/csv; charset=utf-8"
XLSX_CONTENT_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

_FORMULA_PREFIX = re.compile(r"^[\s]*[=+\-@]")
_INVALID_SHEET_CHARS = re.compile(r"[:\\/?*\[\]]")
_INVALID_FILENAME_CHARS = re.compile(r"[^A-Za-z0-9._-]+")


def safe_text(value: str) -> str:
    """Neutralize spreadsheet formulas and invalid XLSX control characters."""

    cleaned = ILLEGAL_CHARACTERS_RE.sub("", value)
    return f"'{cleaned}" if _FORMULA_PREFIX.match(cleaned) else cleaned


def safe_worksheet_name(name: str) -> str:
    cleaned = _INVALID_SHEET_CHARS.sub(" ", name).strip().strip("'")
    return (cleaned or "Report")[:31]


def safe_filename(name: str) -> str:
    cleaned = _INVALID_FILENAME_CHARS.sub("-", name).strip("._-")
    return cleaned or "report"


def _csv_value(value):
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.isoformat(timespec="seconds")
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, str):
        return safe_text(value)
    if isinstance(value, Decimal):
        return str(value)
    return value


def render_csv(report: ReportData) -> bytes:
    """UTF-8 BOM CSV for reliable multilingual spreadsheet compatibility."""

    output = io.StringIO(newline="")
    writer = csv.writer(output, quoting=csv.QUOTE_MINIMAL, lineterminator="\r\n")
    writer.writerow([column.header for column in report.definition.columns])
    for row in report.rows:
        writer.writerow(
            [_csv_value(row.get(column.key)) for column in report.definition.columns]
        )
    return ("\ufeff" + output.getvalue()).encode("utf-8")


def _xlsx_value(value):
    if value is None:
        return None
    if isinstance(value, str):
        return safe_text(value)
    if isinstance(value, Decimal):
        return float(value)
    return value


def _number_format(column: ReportColumn) -> str:
    return {
        "date": "yyyy-mm-dd",
        "datetime": "yyyy-mm-dd hh:mm:ss",
        "integer": "0",
        "decimal": "0.00",
        "percentage": "0.0%",
        "currency": "#,##0.00",
    }.get(column.kind, "General")


def render_xlsx(report: ReportData) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = safe_worksheet_name(report.definition.worksheet_name)

    headers = [column.header for column in report.definition.columns]
    sheet.append(headers)
    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(vertical="top", wrap_text=True)

    for row in report.rows:
        sheet.append(
            [_xlsx_value(row.get(column.key)) for column in report.definition.columns]
        )

    sheet.freeze_panes = "A2"
    last_column = get_column_letter(max(1, len(report.definition.columns)))
    sheet.auto_filter.ref = f"A1:{last_column}{max(1, sheet.max_row)}"

    for index, column in enumerate(report.definition.columns, start=1):
        letter = get_column_letter(index)
        observed = len(column.header)
        for cell in sheet[letter][1:]:
            if cell.value is not None:
                observed = max(observed, len(str(cell.value)))
            cell.number_format = _number_format(column)
            cell.alignment = Alignment(
                vertical="top",
                wrap_text=column.wrap or (isinstance(cell.value, str) and len(cell.value) > 40),
            )
        sheet.column_dimensions[letter].width = min(45, max(10, column.width, observed + 2))

    output = io.BytesIO()
    workbook.save(output)
    workbook.close()
    return output.getvalue()
