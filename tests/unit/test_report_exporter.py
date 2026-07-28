"""CSV/XLSX report serializer correctness and spreadsheet safety."""

import csv
import io
from datetime import date, datetime

from openpyxl import load_workbook

from backend.reports.exporter import render_csv, render_xlsx, safe_worksheet_name
from backend.reports.registry import ReportColumn, ReportData, ReportDefinition


TEST_REPORT = ReportDefinition(
    code="test",
    name="Test",
    worksheet_name="Unsafe:/Name*That Is Far Too Long For Excel Worksheets",
    columns=(
        ReportColumn("name", "Name", "text", 15),
        ReportColumn("note", "Note", "text", 20, wrap=True),
        ReportColumn("date", "Date", "date", 13),
        ReportColumn("created", "Created", "datetime", 20),
        ReportColumn("count", "Count", "integer", 12),
        ReportColumn("amount", "Amount", "currency", 15),
        ReportColumn("rate", "Rate", "percentage", 12),
        ReportColumn("phone", "Phone", "text", 18),
    ),
)


def _report() -> ReportData:
    return ReportData(
        TEST_REPORT,
        [
            {
                "name": "=HYPERLINK(\"https://example.test\")",
                "note": "नमस्ते, \"quoted\"\nsecond line",
                "date": date(2026, 7, 24),
                "created": datetime(2026, 7, 24, 9, 30, 45),
                "count": -7,
                "amount": 1234.5,
                "rate": 0.375,
                "phone": "+919876543210",
            }
        ],
    )


def test_csv_utf8_headers_escaping_types_and_formula_safety():
    content = render_csv(_report())
    assert content.startswith(b"\xef\xbb\xbf")
    rows = list(csv.reader(io.StringIO(content.decode("utf-8-sig"))))

    assert rows[0] == [column.header for column in TEST_REPORT.columns]
    assert rows[1][0].startswith("'=")
    assert rows[1][1] == "नमस्ते, \"quoted\"\nsecond line"
    assert rows[1][2] == "2026-07-24"
    assert rows[1][3] == "2026-07-24T09:30:45"
    assert rows[1][4] == "-7"  # numeric negatives are not treated as formulas
    assert rows[1][7] == "'+919876543210"


def test_xlsx_headers_freeze_filter_widths_types_and_formula_safety():
    workbook = load_workbook(io.BytesIO(render_xlsx(_report())))
    sheet = workbook.active

    assert sheet.title == safe_worksheet_name(TEST_REPORT.worksheet_name)
    assert len(sheet.title) <= 31
    assert sheet.freeze_panes == "A2"
    assert sheet.auto_filter.ref == "A1:H2"
    assert [cell.value for cell in sheet[1]] == [
        column.header for column in TEST_REPORT.columns
    ]
    assert all(cell.font.bold for cell in sheet[1])
    assert sheet["A2"].value.startswith("'=")
    assert sheet["B2"].value == "नमस्ते, \"quoted\"\nsecond line"
    assert sheet["C2"].is_date
    assert sheet["D2"].is_date
    assert sheet["E2"].data_type == "n"
    assert sheet["F2"].data_type == "n"
    assert sheet["G2"].data_type == "n"
    assert sheet["G2"].number_format == "0.0%"
    assert sheet["H2"].data_type == "s"
    assert sheet["H2"].value == "'+919876543210"
    assert all(10 <= sheet.column_dimensions[column].width <= 45 for column in "ABCDEFGH")
    workbook.close()


def test_empty_csv_and_xlsx_remain_valid_with_headers():
    report = ReportData(TEST_REPORT, [])

    rows = list(csv.reader(io.StringIO(render_csv(report).decode("utf-8-sig"))))
    assert rows == [[column.header for column in TEST_REPORT.columns]]

    workbook = load_workbook(io.BytesIO(render_xlsx(report)))
    sheet = workbook.active
    assert sheet.max_row == 1
    assert sheet.freeze_panes == "A2"
    assert all(cell.font.bold for cell in sheet[1])
    workbook.close()


def test_all_dangerous_text_prefixes_are_neutralized_in_both_formats():
    for prefix in ("=", "+", "-", "@", "  ="):
        report = ReportData(
            ReportDefinition(
                code="formula",
                name="Formula",
                worksheet_name="Formula",
                columns=(ReportColumn("value", "Value"),),
            ),
            [{"value": f"{prefix}danger"}],
        )
        csv_rows = list(
            csv.reader(io.StringIO(render_csv(report).decode("utf-8-sig")))
        )
        assert csv_rows[1][0].startswith("'")
        workbook = load_workbook(io.BytesIO(render_xlsx(report)))
        assert workbook.active["A2"].value.startswith("'")
        assert workbook.active["A2"].data_type == "s"
        workbook.close()
