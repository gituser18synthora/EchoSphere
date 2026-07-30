"""Operational export routes: files, filters, authorization and isolation."""

import csv
import io
import uuid
from datetime import datetime

import pytest
from fastapi.testclient import TestClient
from openpyxl import load_workbook
from sqlalchemy import delete, func, select

from backend.core.security import create_access_token
from backend.main import app
from backend.reports.operational import (
    CONVERSATIONS_EXPORT,
    INVOICES_EXPORT,
    OPERATIONAL_EXPORT_REGISTRY,
    SUBSCRIPTIONS_EXPORT,
    TRANSCRIPT_EXPORT,
)
from shared.db.mysql import get_sessionmaker
from shared.models import (
    AuditLog,
    ConversationSession,
    Invoice,
    Plan,
    Subscription,
    Tenant,
    User,
    VoiceBot,
)

pytestmark = pytest.mark.integration
API = "/api/v1"
XLSX_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


@pytest.fixture(scope="module")
def client():
    db = get_sessionmaker()()
    before = set(
        db.scalars(
            select(AuditLog.id).where(
                AuditLog.action.in_(
                    (
                        "data.export",
                        "conversation.transcript.export",
                        "invoice.pdf.download",
                    )
                )
            )
        ).all()
    )
    db.close()
    with TestClient(app) as test_client:
        yield test_client
    db = get_sessionmaker()()
    try:
        created = list(
            db.scalars(
                select(AuditLog.id).where(
                    AuditLog.action.in_(
                        (
                            "data.export",
                            "conversation.transcript.export",
                            "invoice.pdf.download",
                        )
                    ),
                    AuditLog.id.not_in(before),
                )
            ).all()
        )
        if created:
            db.execute(delete(AuditLog).where(AuditLog.id.in_(created)))
            db.commit()
    finally:
        db.close()


def bearer(email: str) -> dict[str, str]:
    db = get_sessionmaker()()
    try:
        user = db.scalar(select(User).where(User.email == email))
        assert user is not None
        token = create_access_token(
            user_id=user.id,
            role=user.role.code,
            tenant_id=user.tenant_id,
        )
        return {"Authorization": f"Bearer {token}"}
    finally:
        db.close()


@pytest.fixture(scope="module")
def super_admin():
    return bearer("alex.rivera@aurexion.com")


@pytest.fixture(scope="module")
def tenant_admin():
    return bearer("priya.sharma@meridianhealth.com")


def csv_rows(response) -> list[list[str]]:
    return list(csv.reader(io.StringIO(response.content.decode("utf-8-sig"))))


def assert_xlsx(response, definition, expected_data_rows: int | None = None):
    assert response.status_code == 200, response.text
    assert response.headers["content-type"].startswith(XLSX_TYPE)
    assert response.headers["content-disposition"].endswith('.xlsx"')
    workbook = load_workbook(io.BytesIO(response.content))
    sheet = workbook.active
    assert [cell.value for cell in sheet[1]] == [
        column.header for column in definition.columns
    ]
    assert all(cell.font.bold for cell in sheet[1])
    assert sheet.freeze_panes == "A2"
    assert sheet.auto_filter.ref is not None
    assert len(sheet.title) <= 31
    if expected_data_rows is not None:
        assert sheet.max_row == expected_data_rows + 1
    workbook.close()


@pytest.mark.parametrize("export_type", ("subscriptions", "invoices"))
def test_platform_operational_exports_support_csv_and_formatted_xlsx(
    client,
    super_admin,
    export_type,
):
    definition = OPERATIONAL_EXPORT_REGISTRY[export_type]
    csv_response = client.get(
        f"{API}/exports/{export_type}?format=csv",
        headers=super_admin,
    )
    assert csv_response.status_code == 200, csv_response.text
    assert csv_response.headers["content-type"].startswith("text/csv")
    assert csv_response.headers["content-disposition"].endswith('.csv"')
    assert csv_rows(csv_response)[0] == [
        column.header for column in definition.columns
    ]

    xlsx_response = client.get(
        f"{API}/exports/{export_type}?format=xlsx",
        headers=super_admin,
    )
    assert_xlsx(xlsx_response, definition)


def test_subscription_export_applies_filters_and_crosses_ui_pagination(
    client,
    super_admin,
):
    suffix = uuid.uuid4().hex[:10]
    tenant_id = f"tn_export_{suffix}"
    tenant_name = f"नमस्ते Export Pagination {suffix}"
    subscription_ids = [f"sub_export_{suffix}_{index:02d}" for index in range(55)]
    db = get_sessionmaker()()
    try:
        plan = db.scalar(
            select(Plan).where(
                Plan.is_deleted.is_(False),
                Plan.status == "active",
            ).order_by(Plan.sort_order.asc())
        )
        assert plan is not None
        db.add(
            Tenant(
                id=tenant_id,
                name=tenant_name,
                domain=f"{suffix}.export.example.test",
                status="active",
            )
        )
        # There is no ORM relationship between these mapped classes, so make
        # the parent row visible before inserting the FK children.
        db.flush()
        db.add_all(
            [
                Subscription(
                    id=subscription_id,
                    tenant_id=tenant_id,
                    plan_id=plan.id,
                    status="active",
                    seats=5,
                    bot_limit=2,
                    minutes_included=1000,
                    mrr=100,
                )
                for subscription_id in subscription_ids
            ]
        )
        db.commit()

        response = client.get(
            f"{API}/exports/subscriptions",
            params={
                "format": "csv",
                "search": tenant_name,
                "status": "active",
                "plan": plan.code,
            },
            headers=super_admin,
        )
        assert response.status_code == 200, response.text
        rows = csv_rows(response)
        assert len(rows) == 56  # header + all 55, beyond the API's default page size 50
        assert [row[0] for row in rows[1:]] == sorted(subscription_ids)
        assert all(row[1] == tenant_name for row in rows[1:])
        assert all(row[4] == "active" for row in rows[1:])

        xlsx_response = client.get(
            f"{API}/exports/subscriptions",
            params={"format": "xlsx", "search": tenant_name},
            headers=super_admin,
        )
        assert_xlsx(xlsx_response, SUBSCRIPTIONS_EXPORT, expected_data_rows=55)
    finally:
        db.rollback()
        db.execute(
            delete(Subscription).where(Subscription.id.in_(subscription_ids))
        )
        db.execute(delete(Tenant).where(Tenant.id == tenant_id))
        db.commit()
        db.close()


@pytest.mark.parametrize("export_type,definition", (
    ("subscriptions", SUBSCRIPTIONS_EXPORT),
    ("invoices", INVOICES_EXPORT),
))
def test_empty_operational_exports_are_valid_csv_and_xlsx(
    client,
    super_admin,
    export_type,
    definition,
):
    search = f"no-match-{uuid.uuid4().hex}"
    csv_response = client.get(
        f"{API}/exports/{export_type}",
        params={"format": "csv", "search": search},
        headers=super_admin,
    )
    assert csv_rows(csv_response) == [
        [column.header for column in definition.columns]
    ]
    xlsx_response = client.get(
        f"{API}/exports/{export_type}",
        params={"format": "xlsx", "search": search},
        headers=super_admin,
    )
    assert_xlsx(xlsx_response, definition, expected_data_rows=0)


def test_invalid_operational_export_requests_and_tenant_billing_access_rejected(
    client,
    super_admin,
    tenant_admin,
):
    assert client.get(f"{API}/exports/subscriptions?format=csv").status_code == 401
    assert client.get(
        f"{API}/exports/not-real?format=csv",
        headers=super_admin,
    ).status_code == 404
    assert client.get(
        f"{API}/exports/subscriptions?format=pdf",
        headers=super_admin,
    ).status_code == 422
    assert client.get(
        f"{API}/exports/subscriptions?format=csv&status=not-real",
        headers=super_admin,
    ).status_code == 422
    assert client.get(
        f"{API}/exports/subscriptions?format=csv&plan=not-real",
        headers=super_admin,
    ).status_code == 422
    assert client.get(
        f"{API}/exports/invoices?format=csv&plan=growth",
        headers=super_admin,
    ).status_code == 422
    assert client.get(
        f"{API}/exports/subscriptions?format=csv",
        headers=tenant_admin,
    ).status_code == 403
    assert client.get(
        f"{API}/exports/invoices?format=xlsx",
        headers=tenant_admin,
    ).status_code == 403


def test_invoice_filter_and_pdf_download(client, super_admin, tenant_admin):
    db = get_sessionmaker()()
    try:
        invoice = db.scalar(
            select(Invoice).where(Invoice.is_deleted.is_(False))
            .order_by(Invoice.issued_at.desc())
        )
        assert invoice is not None
        status = invoice.status
        expected_count = int(
            db.scalar(
                select(func.count())
                .select_from(Invoice)
                .join(Tenant, Tenant.id == Invoice.tenant_id)
                .where(
                    Invoice.is_deleted.is_(False),
                    Tenant.is_deleted.is_(False),
                    Invoice.status == status,
                )
            )
            or 0
        )
    finally:
        db.close()

    response = client.get(
        f"{API}/exports/invoices",
        params={"format": "csv", "status": status},
        headers=super_admin,
    )
    rows = csv_rows(response)
    assert len(rows) == expected_count + 1
    assert all(row[4] == status for row in rows[1:])

    pdf = client.get(f"{API}/invoices/{invoice.id}/pdf", headers=super_admin)
    assert pdf.status_code == 200, pdf.text
    assert pdf.headers["content-type"].startswith("application/pdf")
    assert pdf.headers["content-disposition"].endswith('.pdf"')
    assert pdf.content.startswith(b"%PDF-")
    assert client.get(
        f"{API}/invoices/{invoice.id}/pdf",
        headers=tenant_admin,
    ).status_code == 403


def test_conversation_export_filters_and_enforces_tenant_scope(
    client,
    super_admin,
    tenant_admin,
):
    db = get_sessionmaker()()
    try:
        tenant_user = db.scalar(
            select(User).where(User.email == "priya.sharma@meridianhealth.com")
        )
        assert tenant_user is not None and tenant_user.tenant_id
        tenant_id = tenant_user.tenant_id
        expected_flagged = int(
            db.scalar(
                select(func.count())
                .select_from(ConversationSession)
                .join(VoiceBot, VoiceBot.id == ConversationSession.bot_id)
                .where(
                    ConversationSession.tenant_id == tenant_id,
                    ConversationSession.is_deleted.is_(False),
                    VoiceBot.is_deleted.is_(False),
                    ConversationSession.flagged.is_(True),
                )
            )
            or 0
        )
        foreign_tenant = db.scalar(
            select(Tenant.id).where(
                Tenant.id != tenant_id,
                Tenant.is_deleted.is_(False),
            )
        )
        assert foreign_tenant is not None
    finally:
        db.close()

    response = client.get(
        f"{API}/exports/conversations?format=csv&flagged=true",
        headers=tenant_admin,
    )
    assert response.status_code == 200, response.text
    rows = csv_rows(response)
    assert rows[0] == [
        column.header for column in CONVERSATIONS_EXPORT.columns
    ]
    assert len(rows) == expected_flagged + 1
    assert all(row[14] == "Yes" for row in rows[1:])

    xlsx = client.get(
        f"{API}/exports/conversations?format=xlsx&sentiment=negative",
        headers=tenant_admin,
    )
    assert_xlsx(xlsx, CONVERSATIONS_EXPORT)

    assert client.get(
        f"{API}/exports/conversations?format=csv&tenantId={foreign_tenant}",
        headers=tenant_admin,
    ).status_code == 403
    assert client.get(
        f"{API}/exports/conversations?format=csv",
        headers=super_admin,
    ).status_code == 422
    assert client.get(
        f"{API}/exports/conversations?format=csv&tenantId={tenant_id}",
        headers=super_admin,
    ).status_code == 200


def test_conversation_export_crosses_default_pagination_and_empty_is_valid(
    client,
    tenant_admin,
):
    suffix = uuid.uuid4().hex[:10]
    conversation_ids = [f"cv_export_{suffix}_{index:02d}" for index in range(55)]
    db = get_sessionmaker()()
    try:
        tenant_user = db.scalar(
            select(User).where(User.email == "priya.sharma@meridianhealth.com")
        )
        assert tenant_user is not None and tenant_user.tenant_id
        bot = db.scalar(
            select(VoiceBot).where(
                VoiceBot.tenant_id == tenant_user.tenant_id,
                VoiceBot.is_deleted.is_(False),
            )
        )
        assert bot is not None
        db.add_all(
            [
                ConversationSession(
                    id=conversation_id,
                    tenant_id=tenant_user.tenant_id,
                    bot_id=bot.id,
                    channel="voice",
                    caller_masked="•••",
                    started_at=datetime(2026, 7, 24, 12, 0, 0),
                    duration_sec=index,
                    sentiment="neutral",
                    intents=["pagination_test"],
                    contained=True,
                    cost_usd=0,
                    language="en-IN",
                    flagged=False,
                    status="completed",
                )
                for index, conversation_id in enumerate(conversation_ids)
            ]
        )
        db.commit()

        response = client.get(
            f"{API}/exports/conversations",
            params={"format": "csv", "search": f"cv_export_{suffix}"},
            headers=tenant_admin,
        )
        rows = csv_rows(response)
        assert len(rows) == 56
        assert {row[0] for row in rows[1:]} == set(conversation_ids)

        no_match = f"no-conversation-{uuid.uuid4().hex}"
        empty_csv = client.get(
            f"{API}/exports/conversations",
            params={"format": "csv", "search": no_match},
            headers=tenant_admin,
        )
        assert csv_rows(empty_csv) == [
            [column.header for column in CONVERSATIONS_EXPORT.columns]
        ]
        empty_xlsx = client.get(
            f"{API}/exports/conversations",
            params={"format": "xlsx", "search": no_match},
            headers=tenant_admin,
        )
        assert_xlsx(empty_xlsx, CONVERSATIONS_EXPORT, expected_data_rows=0)
    finally:
        db.rollback()
        db.execute(
            delete(ConversationSession).where(
                ConversationSession.id.in_(conversation_ids)
            )
        )
        db.commit()
        db.close()


def test_transcript_export_csv_xlsx_and_tenant_isolation(
    client,
    tenant_admin,
    super_admin,
):
    db = get_sessionmaker()()
    try:
        tenant_user = db.scalar(
            select(User).where(User.email == "priya.sharma@meridianhealth.com")
        )
        assert tenant_user is not None and tenant_user.tenant_id
        own_conversation = db.scalar(
            select(ConversationSession)
            .where(
                ConversationSession.tenant_id == tenant_user.tenant_id,
                ConversationSession.is_deleted.is_(False),
            )
            .order_by(ConversationSession.started_at.desc())
        )
        foreign_conversation = db.scalar(
            select(ConversationSession).where(
                ConversationSession.tenant_id != tenant_user.tenant_id,
                ConversationSession.is_deleted.is_(False),
            )
        )
        assert own_conversation is not None
        assert foreign_conversation is not None
    finally:
        db.close()

    csv_response = client.get(
        f"{API}/conversations/{own_conversation.id}/transcript/export?format=csv",
        headers=tenant_admin,
    )
    assert csv_response.status_code == 200, csv_response.text
    assert csv_rows(csv_response)[0] == [
        column.header for column in TRANSCRIPT_EXPORT.columns
    ]
    xlsx_response = client.get(
        f"{API}/conversations/{own_conversation.id}/transcript/export?format=xlsx",
        headers=tenant_admin,
    )
    assert_xlsx(xlsx_response, TRANSCRIPT_EXPORT)

    assert client.get(
        f"{API}/conversations/{foreign_conversation.id}/transcript/export?format=csv",
        headers=tenant_admin,
    ).status_code == 404
    assert client.get(
        f"{API}/conversations/{foreign_conversation.id}/transcript/export?format=csv",
        headers=super_admin,
    ).status_code == 200


def test_operational_exports_create_high_level_audit_records(
    client,
    super_admin,
):
    db = get_sessionmaker()()
    before = set(
        db.scalars(
            select(AuditLog.id).where(
                AuditLog.action == "data.export",
                AuditLog.entity_id == "subscriptions",
            )
        ).all()
    )
    db.close()
    response = client.get(
        f"{API}/exports/subscriptions?format=csv&status=active",
        headers=super_admin,
    )
    assert response.status_code == 200
    db = get_sessionmaker()()
    try:
        audit = db.scalar(
            select(AuditLog)
            .where(
                AuditLog.action == "data.export",
                AuditLog.entity_id == "subscriptions",
                AuditLog.id.not_in(before),
            )
            .order_by(AuditLog.created_at.desc())
        )
        assert audit is not None
        assert audit.new_value["format"] == "csv"
        assert audit.new_value["filters"]["status"] == "active"
        assert audit.new_value["filters"]["hasSearch"] is False
        assert "password" not in str(audit.new_value).lower()
        assert "token" not in str(audit.new_value).lower()
    finally:
        db.close()
